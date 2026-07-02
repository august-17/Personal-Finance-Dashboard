import csv
from datetime import datetime
from tkinter import filedialog, messagebox

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from constants import CSV_HEADERS

from finance import get_monthly_transactions

from dialogs import (
    open_month_selector, 
    open_export_window
)



def export_selected_format(format_combobox, export_window, selected_month):

    selected_format = format_combobox.get()

    export_window.destroy()

    export_functions = {

        "CSV": export_csv,

        "PDF": export_pdf,

        "Excel": export_excel

    }

    export_functions[selected_format](selected_month)


def export_report(root):

    open_month_selector(
        root,
        "Export Report",
        lambda selector, month, year: 
            open_export_window(root, selector, month, year, export_selected_format)
    )


def get_export_transactions(selected_month):

    transactions = get_monthly_transactions(selected_month)

    if not transactions:

        messagebox.showinfo(
            "No Data",
            "No transactions available for the selected month."
        )

        return None

    return transactions


def export_csv(selected_month):

    file_path = filedialog.asksaveasfilename(
        defaultextension=".csv",
        filetypes=[("CSV Files", "*.csv")],
        title="Save Financial Report"
    )

    if not file_path:

        return
    
    transactions = get_export_transactions(selected_month)

    if transactions is None:

        return

    try:

        with open(file_path, "w", newline="", encoding="utf-8") as file:

            writer = csv.writer(file)

            writer.writerow(CSV_HEADERS)

            for row in transactions:

                writer.writerow([
                    row["ID"],
                    row["Date"],
                    row["Type"],
                    row["Category"],
                    row["Amount"],
                    row["Description"]
                ])

        messagebox.showinfo(
            "Success",
            "Financial report exported successfully."
        )

    except Exception as e:

        messagebox.showerror(
            "Export Error",
            f"Unable to export report.\n\n{e}"
        )


def export_pdf(selected_month):

    file_path = filedialog.asksaveasfilename(
        defaultextension=".pdf",
        filetypes=[("PDF Files", "*.pdf")],
        title="Save Financial Report"
    )

    if not file_path:

        return
    
    transactions = get_export_transactions(selected_month)

    if transactions is None:

        return
    
    pdf = SimpleDocTemplate(file_path, pagesize=landscape(A4))

    styles = getSampleStyleSheet()

    title_style = styles["Heading1"]

    normal_style = styles["Normal"]
    
    table_data = [["Date", "Type", "Category", "Amount", "Description"]]

    for row in transactions:

        table_data.append([
            row["Date"],
            row["Type"],
            row["Category"],
            f"Rs.{float(row['Amount']):,.2f}",
            row["Description"]
        ])

    month_display = datetime.strptime(selected_month, "%Y-%m").strftime("%B %Y")

    title = Paragraph(f"Personal Finance Report ({month_display})", title_style)

    generated_on = Paragraph(
        f"Generated On: {datetime.now().strftime('%d %B %Y, %I:%M %p')}",
        normal_style
    )

    table = Table(table_data, colWidths=[90, 70, 100, 80, 240], repeatRows=1)

    table.setStyle(

        TableStyle([

            ("BACKGROUND", (0, 0), (-1, 0), colors.grey),

            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),

            ("ALIGN", (0, 0), (-1, -1), "CENTER"),

            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),

            ("BOTTOMPADDING", (0, 0), (-1, 0), 10),

            ("GRID", (0, 0), (-1, -1), 1, colors.black),

            ("BACKGROUND", (0, 1), (-1, -1), colors.beige)

        ])

    )

    try:

        table.splitByRow = True

        pdf.build([title, Spacer(1, 12), generated_on, Spacer(1, 20), table])

        messagebox.showinfo(
            "Success",
            "Financial report exported successfully."
        )

    except Exception as e:

        messagebox.showerror(
            "Export Error",
            f"Unable to export report.\n\n{e}"
        )


def export_excel(selected_month):

    file_path = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Excel Files", "*.xlsx")],
        title="Save Financial Report"
    )

    if not file_path:

        return
    
    transactions = get_export_transactions(selected_month)

    if transactions is None:

        return

    try:

        workbook = Workbook()

        worksheet = workbook.active

        worksheet.title = "Financial Report"

        worksheet.append(CSV_HEADERS)

        for cell in worksheet[1]:

            cell.font = Font(bold=True)

        for row in transactions:

            worksheet.append([
                row["ID"],
                row["Date"],
                row["Type"],
                row["Category"],
                float(row["Amount"]),
                row["Description"]
            ])

        # Format Amount Column
        for cell in worksheet["E"][1:]:

            cell.number_format = '#,##0.00'

        # Auto-fit Columns
        for column_cells in worksheet.columns:

            max_length = 0

            column = get_column_letter(column_cells[0].column)

            for cell in column_cells:

                try:

                    max_length = max(max_length, len(str(cell.value)))

                except Exception:

                    pass

            worksheet.column_dimensions[column].width = max_length + 2
        
        workbook.save(file_path)

        messagebox.showinfo(
            "Success",
            "Financial report exported successfully."
        )

    except Exception as e:

        messagebox.showerror(
            "Export Error",
            f"Unable to export report.\n\n{e}"
        )