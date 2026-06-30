import tkinter as tk
from tkinter import ttk, messagebox
from datetime import datetime, timedelta
from tkcalendar import DateEntry
from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer)
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
import re
import csv
import json
import os
from tkinter import filedialog
import shutil
import matplotlib.pyplot as plt

CATEGORIES = ["Food", "Travel", "Shopping", "Bills", "Education", "Healthcare", "Entertainment", "Other"]

EXPORT_FORMATS = ["CSV", "PDF", "Excel"]

CSV_HEADERS = ["ID", "Date", "Type", "Category", "Amount", "Description"]

CSV_FILE = os.path.join(os.path.dirname(__file__), "transactions.csv")

BUDGET_FILE = os.path.join(os.path.dirname(__file__), "budget.txt")

CATEGORY_BUDGET_FILE = os.path.join(os.path.dirname(__file__), "category_budget.json")

BACKUP_FILE = os.path.join(os.path.dirname(__file__), "backup_transactions.csv")

MIN_AMOUNT = 1
MAX_AMOUNT = 10000000

MAX_DESCRIPTION_LENGTH = 50

MIN_PIE_PERCENTAGE = 2

LABEL_FONT = ("Arial", 12, "bold")
TITLE_FONT = ("Arial", 18, "bold")
REPORT_FONT = ("Courier New", 11)

WINDOW_WIDTH = 1300
WINDOW_HEIGHT = 850

REPORT_WIDTH = 850
REPORT_HEIGHT = 450

BUDGET_WIDTH = 500
BUDGET_HEIGHT = 550

EXPORT_WIDTH = 300
EXPORT_HEIGHT = 150

SELECTOR_WIDTH = 300
SELECTOR_HEIGHT = 180

EVEN_ROW_COLOR = "#f5f5f5"
ODD_ROW_COLOR = "white"

editing_transaction_id = None

sort_reverse = {"Date": True}
sort_column = "Date"

delete_history = []

search_after_id = None



def handle_category_change(event):

    if category_combobox.get() == "Other":

        custom_category_label.grid(row=5, column=0, padx=5, pady=5)
        custom_category_entry.grid(row=5, column=1, padx=5, pady=5)

    else:

        custom_category_label.grid_remove()
        custom_category_entry.grid_remove()



def clear_inputs():

    amount_entry.delete(0, tk.END)

    description_entry.delete(0, tk.END)

    custom_category_entry.delete(0, tk.END)

    category_combobox.current(0)

    custom_category_label.grid_remove()
    custom_category_entry.grid_remove()

    date_entry.set_date(datetime.now())

    type_combobox.current(0)



def enable_mousewheel_scrolling(widget):

    def _on_mousewheel(event):

        widget.yview_scroll(int(-1 * (event.delta / 120)), "units")

    widget.bind("<Enter>", lambda e: widget.bind_all("<MouseWheel>", _on_mousewheel))

    widget.bind("<Leave>", lambda e: widget.unbind_all("<MouseWheel>"))



def create_csv_file():

    if not os.path.exists(CSV_FILE):

        with open(CSV_FILE, "w", newline="", encoding="utf-8") as file:

            writer = csv.writer(file)

            writer.writerow(CSV_HEADERS)



def create_backup():

    if os.path.exists(CSV_FILE):

        shutil.copy(CSV_FILE, BACKUP_FILE)



def read_transactions():

    with open(CSV_FILE, "r", newline="", encoding="utf-8") as file:

        return list(csv.DictReader(file))



def get_all_categories():

    categories = set(CATEGORIES)

    transactions = read_transactions()

    for row in transactions:
        categories.add(row["Category"])

    return sorted(categories)



def load_budget():

    if not os.path.exists(BUDGET_FILE):

        return 0

    try:

        with open(BUDGET_FILE, "r", encoding="utf-8") as file:

            return float(file.read())

    except ValueError:

        return 0


def check_budget_warning(amount):

    if type_combobox.get() != "Expense":

        return True

    budget = load_budget()

    if budget == 0:

        return True

    current_month = datetime.now().strftime("%Y-%m")

    monthly_expenses = 0

    transactions = read_transactions()

    for row in transactions:

        if (
            row["Type"] == "Expense"
            and row["Date"].startswith(current_month)
        ):
            
            monthly_expenses += float(row["Amount"])

    if monthly_expenses + amount > budget:

        return messagebox.askyesno(
            "Budget Warning",
            "This transaction exceeds your monthly budget.\n\nDo you want to continue?"
        )

    return True



def check_category_budget_warning(category, amount):

    if type_combobox.get() != "Expense":

        return True

    category_budgets = load_category_budgets()

    if category not in category_budgets:

        return True

    budget = category_budgets[category]

    current_month = datetime.now().strftime("%Y-%m")

    category_spent = 0

    transactions = read_transactions()

    for row in transactions:

        if (
            row["Type"] == "Expense"
            and row["Category"] == category
            and row["Date"].startswith(current_month)
        ):

            category_spent += float(row["Amount"])

    if category_spent + amount > budget:

        return messagebox.askyesno(
            "Category Budget Warning",
            f"This transaction exceeds the budget for '{category}'.\n\n"
            "Do you want to continue?"
        )

    return True



def load_category_budgets():

    if not os.path.exists(CATEGORY_BUDGET_FILE):

        return {}

    try:

        with open(CATEGORY_BUDGET_FILE, "r", encoding="utf-8") as file:

            return json.load(file)

    except (FileNotFoundError, json.JSONDecodeError):

        return {}



def save_category_budgets(category_budgets):

    with open(CATEGORY_BUDGET_FILE, "w", encoding="utf-8") as file:

        json.dump(category_budgets, file, indent=4)



def open_category_budget_window():

    budget_window = tk.Toplevel(root)

    budget_window.title("Category Budgets")

    budget_window.geometry(f"{BUDGET_WIDTH}x{BUDGET_HEIGHT}")

    canvas = tk.Canvas(budget_window)

    scrollbar = tk.Scrollbar(
        budget_window,
        orient="vertical",
        command=canvas.yview
    )

    scrollable_frame = tk.Frame(canvas)

    scrollable_frame.bind(
        "<Configure>",
        lambda e: canvas.configure(
            scrollregion=canvas.bbox("all")
        )
    )

    canvas.create_window(
        (0, 0),
        window=scrollable_frame,
        anchor="nw"
    )

    canvas.configure(yscrollcommand=scrollbar.set)

    enable_mousewheel_scrolling(canvas)

    canvas.pack(side="left", fill="both", expand=True)

    scrollbar.pack(side="right", fill="y")

    saved_budgets = load_category_budgets()

    category_entries = {}

    all_categories = get_all_categories()

    for index, category in enumerate(all_categories):

        tk.Label(scrollable_frame, text=category).grid(row=index, column=0, padx=(15, 10), pady=6, sticky="w")

        entry = tk.Entry(scrollable_frame, width=15)

        entry.grid(row=index, column=1, padx=(15, 10), pady=6)

        reset_button = tk.Button(
            scrollable_frame,
            text="Reset",
            width=8,
            command=lambda e=entry: reset_single_category(e)
        )

        reset_button.grid(row=index, column=2, padx=5, pady=6)

        if category in saved_budgets:

            entry.insert(
                0,
                str(saved_budgets[category])
            )

        category_entries[category] = entry

    save_button = tk.Button(
        scrollable_frame,
        text="Save Budgets",
        command=lambda: save_category_budget_entries(
            category_entries,
            budget_window
        )
    )

    save_button.grid(row=len(all_categories), column=0, columnspan=3, padx=10, pady=15)



def reset_single_category(entry):

    entry.delete(0, tk.END)



def save_category_budget_entries(category_entries, budget_window):

    category_budgets = {}

    for category, entry in category_entries.items():

        budget = entry.get().strip()

        if not budget:

            continue

        try:

            budget = float(budget)

        except ValueError:

            messagebox.showerror(
                "Error",
                f"Budget for {category} must be a number."
            )

            return

        if budget <= 0:

            messagebox.showerror(
                "Error",
                f"Budget for {category} must be greater than zero."
            )

            return

        if budget > MAX_AMOUNT:

            messagebox.showerror(
                "Error",
                f"Budget for {category} cannot exceed ₹{MAX_AMOUNT:,}."
            )

            return

        category_budgets[category] = budget

    save_category_budgets(category_budgets)

    messagebox.showinfo(
        "Success",
        "Category budgets saved successfully."
    )

    budget_window.destroy()



def save_budget():

    budget = budget_entry.get().strip()

    if not budget:

        messagebox.showerror(
            "Error",
            "Please enter a budget amount."
        )
        return

    try:

        budget = float(budget)

    except ValueError:

        messagebox.showerror(
            "Error",
            "Budget must be a number."
        )
        return

    if budget <= 0:

        messagebox.showerror(
            "Error",
            "Budget must be greater than zero."
        )
        return
    
    if budget > MAX_AMOUNT:

        messagebox.showerror(
            "Error",
            f"Budget cannot exceed ₹{MAX_AMOUNT:,}."
        )
        return

    with open(BUDGET_FILE, "w", encoding="utf-8") as file:

        file.write(str(budget))

    update_summary()

    messagebox.showinfo(
        "Success",
        "Budget saved successfully."
    )



def reset_budget():

    confirm = messagebox.askyesno(
        "Reset Budget",
        "Are you sure you want to reset the budget?"
    )

    if not confirm:

        return

    if os.path.exists(BUDGET_FILE):

        os.remove(BUDGET_FILE)

    budget_entry.delete(0, tk.END)

    update_summary()

    messagebox.showinfo(
        "Success",
        "Budget reset successfully."
    )



def get_next_id():

    highest_id = 0

    transactions = read_transactions()

    for row in transactions:

        current_id = int(row["ID"])

        if current_id > highest_id:

            highest_id = current_id

    return highest_id + 1



def validate_amount():

    amount = amount_entry.get().strip()

    if not amount:

        messagebox.showerror(
            "Error",
            "Please enter an amount."
        )

        return None

    try:

        amount = float(amount)

    except ValueError:

        messagebox.showerror(
            "Error",
            "Amount must be a number."
        )

        return None

    if amount < MIN_AMOUNT:

        messagebox.showerror(
            "Error",
            f"Amount must be at least ₹{MIN_AMOUNT}."
        )

        return None

    if amount > MAX_AMOUNT:

        messagebox.showerror(
            "Error",
            f"Amount cannot exceed ₹{MAX_AMOUNT:,}."
        )

        return None

    return amount



def validate_description():

    description = description_entry.get().strip()

    if len(description) > MAX_DESCRIPTION_LENGTH:

        messagebox.showerror(
            "Error",
            f"Description cannot exceed {MAX_DESCRIPTION_LENGTH} characters."
        )

        return None

    return description



def get_category():

    category = category_combobox.get()

    if category != "Other":

        return category

    category = custom_category_entry.get().strip()

    if not category:

        messagebox.showerror(
            "Error",
            "Please enter a custom category."
        )

        return None
    
    if len(category) > 30:

        messagebox.showerror(
            "Error",
            "Category name cannot exceed 30 characters."
        )

        return None

    if not re.search(r"[A-Za-z]", category):

        messagebox.showerror(
            "Error",
            "Category must contain at least one alphabet."
        )

        return None

    if not re.fullmatch(r"[A-Za-z0-9 &()-]+", category):

        messagebox.showerror(
            "Error",
            "Category contains invalid characters."
        )

        return None


    return category.title()



def add_transaction():

    transaction_type = type_combobox.get()

    category = get_category()

    if category is None:

        return

    amount = validate_amount()

    if amount is None:

        return
    
    if not check_budget_warning(amount):

        return
    
    if not check_category_budget_warning(category, amount):

        return
    
    description = validate_description()

    if description is None:

        return
    
    transaction_id = get_next_id()

    date = date_entry.get()

    create_backup()

    with open(CSV_FILE, "a", newline="", encoding="utf-8") as file:

        writer = csv.writer(file)

        writer.writerow([transaction_id, date, transaction_type, category, amount, description])

    clear_inputs()

    update_summary()

    apply_filter()

    



def insert_row(values):

    row_count = len(tree.get_children())

    tag = "evenrow" if row_count % 2 == 0 else "oddrow"

    tree.insert("", "end", values=values, tags=(tag,))



def load_transactions():

    try:

        transactions = read_transactions()

        transactions.sort(key=lambda row: row["Date"], reverse=True)

        for row in transactions:

            insert_row((
                row["ID"],
                row["Date"],
                row["Type"],
                row["Category"],
                row["Amount"],
                row["Description"]
            ))

    except Exception as e:

        messagebox.showerror(
            "CSV Error",
            f"Unable to load transactions.\n\n{e}"
        )



def update_summary():

    total_income = 0
    total_expenses = 0
    monthly_expenses = 0

    try:

        budget = load_budget()

        current_month = datetime.now().strftime("%Y-%m")

        transactions = read_transactions()

        for row in transactions:

            amount = float(row["Amount"])

            if row["Type"] == "Income":

                total_income += amount

            else:

                total_expenses += amount

                if row["Date"].startswith(current_month):

                    monthly_expenses += amount

        balance = total_income - total_expenses

        remaining_budget = budget - monthly_expenses

        income_label.config(text=f"Total Income: ₹{total_income:,.2f}")

        expense_label.config(text=f"Total Expenses: ₹{total_expenses:,.2f}")

        if balance >= 0:

            balance_label.config(text=f"Current Balance: ₹{balance:,.2f}", fg="green")

        else:

            balance_label.config(text=f"Current Deficit: ₹{abs(balance):,.2f}", fg="red")

        budget_label.config(text=f"Monthly Budget: ₹{budget:,.2f}")

        if budget == 0:

            status_label.config(text="Budget Status: Not Set", fg="black")

        elif remaining_budget >= 0:

            status_label.config(text=f"Budget Status: ₹{remaining_budget:,.2f} Remaining", fg="green")

        else:

            status_label.config(text=f"Budget Status: ₹{abs(remaining_budget):,.2f} Exceeded", fg="red")

    except Exception as e:

        messagebox.showerror(
            "CSV Error",
            f"Unable to calculate summary.\n\n{e}"
        )



def apply_filter():

    for item in tree.get_children():

        tree.delete(item)

    search_text = " ".join(search_entry.get().lower().strip())

    selected_type = filter_combobox.get()

    transactions = read_transactions()

    for row in transactions:

        date = " ".join(row["Date"].lower().split())
        transaction_type = " ".join(row["Type"].lower().split())
        category = " ".join(row["Category"].lower().split())
        amount = " ".join(row["Amount"].lower().split())
        description = " ".join(row["Description"].lower().split())

        matches_search = (
            search_text in date
            or search_text in transaction_type
            or search_text in category
            or search_text in amount
            or search_text in description
        )

        matches_type = (selected_type == "All" or row["Type"] == selected_type)

        if matches_search and matches_type:
                
            insert_row((
                row["ID"],
                row["Date"],
                row["Type"],
                row["Category"],
                row["Amount"],
                row["Description"]
            ))

    if sort_column is not None:

        sort_treeview(sort_column,toggle=False)



def debounce_search(event=None):

    global search_after_id

    if search_after_id is not None:

        root.after_cancel(search_after_id)

    search_after_id = root.after(250, apply_filter)



def reset_filter():

    search_entry.delete(0, tk.END)

    filter_combobox.current(0)

    apply_filter()



def update_sort_headers():

    sortable_columns = ("Date", "Amount", "Category")

    for column in columns:

        heading = column

        if column == sort_column:

            if sort_reverse.get(column, False):

                heading = f"{column} ▼"

            else:

                heading = f"{column} ▲"

        if column in sortable_columns:

            tree.heading(
                column,
                text=heading,
                command=lambda c=column: sort_treeview(c)
            )

        else:

            tree.heading(
                column,
                text=heading
            )



def sort_treeview(column, toggle=True):

    global sort_column

    if toggle:

        if sort_column == column:

            sort_reverse[column] = not sort_reverse.get(column, False)

        else:

            sort_column = column

            sort_reverse.setdefault(column, False)

    else:

        sort_column = column

    data = []

    for item in tree.get_children():

        values = tree.item(item)["values"]

        data.append((values, item))

    if column == "Amount":

        data.sort(
            key=lambda x: float(x[0][4]),
            reverse=sort_reverse.get(column, False)
        )

    elif column == "Date":

        data.sort(
            key=lambda x: datetime.strptime(x[0][1], "%Y-%m-%d"),
            reverse=sort_reverse.get(column, False)
        )

    elif column == "Category":

        data.sort(
            key=lambda x: x[0][3].lower(),
            reverse=sort_reverse.get(column, False)
        )

    else:

        return

    for index, (_, item) in enumerate(data):

        tree.move(item, "", index)

    update_sort_headers()



def delete_transaction():

    global delete_history

    selected_item = tree.selection()

    if not selected_item:

        messagebox.showwarning(
            "No Selection",
            "Please select a transaction to delete."
        )
        return

    confirm = messagebox.askyesno(
        "Confirm Delete",
        f"Delete {len(selected_item)} selected transaction(s)?"
    )

    if not confirm:
        return

    transaction_ids = set()

    current_delete = []

    for item in selected_item:

        values = tree.item(item, "values")
        transaction_ids.add(values[0])

    transactions = read_transactions()

    updated_transactions = []

    for index, transaction in enumerate(transactions):

        if transaction["ID"] in transaction_ids:
                
            current_delete.append((index, transaction.copy()))
                
            continue

        updated_transactions.append(transaction)

    create_backup()

    with open(CSV_FILE, "w", newline="", encoding="utf-8") as file:

        writer = csv.DictWriter(file, fieldnames=CSV_HEADERS)

        writer.writeheader()

        writer.writerows(updated_transactions)

    delete_history.append(current_delete)

    apply_filter()

    update_summary()

    undo_delete_button.config(state="normal")



def undo_delete():

    global delete_history

    if not delete_history:

        messagebox.showinfo(
            "Undo Delete",
            "Nothing to undo."
        )

        return
    
    last_deleted_transactions = delete_history.pop()
    
    transactions = read_transactions()

    for index, transaction in sorted(last_deleted_transactions, key=lambda x: x[0]):

        transactions.insert(index, transaction)

    create_backup()

    with open(CSV_FILE, "w", newline="", encoding="utf-8") as file:

        writer = csv.DictWriter(file, fieldnames=CSV_HEADERS)

        writer.writeheader()

        writer.writerows(transactions)

    if not delete_history:

        undo_delete_button.config(state="disabled")

    apply_filter()

    update_summary()

    messagebox.showinfo(
        "Undo Delete",
        "Deleted transaction(s) restored successfully."
    )



def edit_transaction():

    global editing_transaction_id

    selected_item = tree.selection()

    if len(selected_item) > 1:

        messagebox.showwarning(
            "Multiple Selection",
            "Please select only one transaction to edit."
        )
        return

    if not selected_item:

        messagebox.showwarning(
            "No Selection",
            "Please select a transaction to edit."
        )
        return

    values = tree.item(selected_item[0], "values")

    editing_transaction_id = values[0]

    date_entry.set_date(values[1])

    type_combobox.set(values[2])

    category = values[3]

    if category in CATEGORIES:

        category_combobox.set(category)

        custom_category_label.grid_remove()
        custom_category_entry.grid_remove()

    else:

        category_combobox.set("Other")

        custom_category_label.grid(row=5, column=0, padx=5, pady=5)
        custom_category_entry.grid(row=5, column=1, padx=5,pady=5)

        custom_category_entry.delete(0, tk.END)
        custom_category_entry.insert(0, category)

    amount_entry.delete(0, tk.END)
    amount_entry.insert(0, values[4])

    description_entry.delete(0, tk.END)
    description_entry.insert(0, values[5])



def save_changes():

    global editing_transaction_id

    if editing_transaction_id is None:

        messagebox.showwarning(
            "No Edit",
            "Select a transaction to edit first."
        )
        return

    transaction_type = type_combobox.get()

    date = date_entry.get()

    category = get_category()

    if category is None:

        return

    amount = validate_amount()

    if amount is None:

        return
    
    description = validate_description()

    if description is None:

        return
    
    rows = []

    with open(CSV_FILE, "r", newline="", encoding="utf-8") as file:

        reader = csv.reader(file)

        header = next(reader)

        for row in reader:

            if row[0] == str(editing_transaction_id):

                row = [row[0], date, transaction_type, category, amount, description]

            rows.append(row)

    create_backup()

    with open(CSV_FILE, "w", newline="", encoding="utf-8") as file:

        writer = csv.writer(file)

        writer.writerow(header)

        writer.writerows(rows)

    editing_transaction_id = None

    apply_filter()

    update_summary()

    clear_inputs()

    messagebox.showinfo(
        "Success",
        "Transaction updated successfully."
    )



def get_available_years():

    years = set()

    transactions = read_transactions()

    for row in transactions:

        years.add(row["Date"][:4])

    years.add(str(datetime.now().year))

    return sorted(years)



def open_month_selector(button_text, callback):

    selector = tk.Toplevel(root)

    selector.title("Select Month")

    selector.geometry(f"{SELECTOR_WIDTH}x{SELECTOR_HEIGHT}")

    selector.resizable(False, False)

    tk.Label(selector, text="Month").pack(pady=(15, 5))

    months = [
        "January", "February", "March", "April",
        "May", "June", "July", "August",
        "September", "October", "November", "December"
    ]

    month_combobox = ttk.Combobox(
        selector,
        values=months,
        state="readonly",
        width=20
    )

    month_combobox.current(datetime.now().month - 1)

    month_combobox.pack()

    tk.Label(selector, text="Year").pack(pady=(10, 5))

    available_years = get_available_years()

    year_combobox = ttk.Combobox(
        selector,
        values=available_years,
        state="readonly",
        width=20
    )

    year_combobox.set(str(datetime.now().year))

    year_combobox.pack()

    tk.Button(
        selector,
        text=button_text,
        command=lambda: callback(
            selector,
            month_combobox.get(),
            year_combobox.get()
        )
    ).pack(pady=15)



def show_monthly_expense_breakdown(selected_month):

    category_totals = {}

    transactions = read_transactions()

    for row in transactions:

        if row["Type"] == "Expense" and row["Date"].startswith(selected_month):

            category = row["Category"]
            amount = float(row["Amount"])

            if category not in category_totals:

                category_totals[category] = 0

            category_totals[category] += amount

    if not category_totals:

        messagebox.showinfo(
            "No Data",
            "No expense data available."
        )
        return
    
    sorted_categories = sorted(
        category_totals.items(),
        key=lambda item: item[1],
        reverse=True
    )

    labels = [category for category, amount in sorted_categories]

    amounts = [amount for category, amount in sorted_categories]

    total_expense = sum(amounts)

    filtered_labels = []
    filtered_amounts = []

    others_total = 0

    for label, amount in zip(labels, amounts):

        percentage = (amount / total_expense) * 100

        if percentage < MIN_PIE_PERCENTAGE:

            others_total += amount

        else:

            filtered_labels.append(label)
            filtered_amounts.append(amount)

    if others_total > 0:

        filtered_labels.append("Others")

        filtered_amounts.append(others_total)

    plt.figure(figsize=(10, 7))

    wedges, _, _ = plt.pie(
        filtered_amounts,
        autopct="%1.1f%%",
        startangle=90
    )

    plt.legend(
        wedges,
        filtered_labels,
        title="Categories",
        loc="center left",
        bbox_to_anchor=(1, 0.5)
    )

    display_month = datetime.strptime(selected_month, "%Y-%m").strftime("%B %Y")

    plt.title(f"Expense Breakdown by Category ({display_month})")

    plt.show()



def generate_expense_breakdown(selector, month_name, year):

    month_number = datetime.strptime(month_name, "%B").month

    selected_month = f"{year}-{month_number:02d}"

    selector.destroy()

    show_monthly_expense_breakdown(selected_month)



def show_monthly_category_report(selected_month):

    category_totals = {}

    transactions = read_transactions()

    for row in transactions:

        if (row["Type"] == "Expense" and row["Date"].startswith(selected_month)):

            category = row["Category"]

            amount = float(row["Amount"])

            category_totals[category] = (category_totals.get(category, 0) + amount  )

    if not category_totals:

        messagebox.showinfo(
            "No Data",
            "No expense data available."
        )
        return
    
    display_month = datetime.strptime(selected_month, "%Y-%m").strftime("%B %Y")

    report_window = tk.Toplevel(root)

    report_window.title("Category-Wise Spending Report")

    report_window.geometry(f"{REPORT_WIDTH}x{REPORT_HEIGHT}")

    report_window.resizable(False, False)

    scrollbar = tk.Scrollbar(report_window)

    scrollbar.pack(side="right", fill="y")

    text = tk.Text(report_window, font=REPORT_FONT, yscrollcommand=scrollbar.set)

    text.pack(fill="both", expand=True, padx=10, pady=10)

    scrollbar.config(command=text.yview)

    enable_mousewheel_scrolling(text)

    report = f"Category Report ({display_month})\n"

    report += "=" * 40 + "\n\n"

    report += f"{'Category':<25}{'Amount'}\n"

    report += "-" * 40 + "\n"

    total = 0

    for category, amount in sorted(
        category_totals.items(),
        key=lambda item: item[1],
        reverse=True
    ):

        report += (
            f"{category:<25}"
            f"₹{amount:,.2f}\n"
        )

        total += amount

    report += "-" * 40 + "\n"

    report += (
        f"{'Total Expenses':<25}"
        f"₹{total:,.2f}"
    )

    text.insert("1.0", report)

    text.config(state="disabled")



def generate_category_report(selector, month_name, year):

    month_number = datetime.strptime(month_name,"%B").month

    selected_month = f"{year}-{month_number:02d}"

    selector.destroy()

    show_monthly_category_report(selected_month)



def show_monthly_summary(selected_month):

    monthly_income = 0
    monthly_expenses = 0

    income_count = 0
    expense_count = 0

    transactions = read_transactions()

    for row in transactions:

        if not row["Date"].startswith(selected_month):

            continue

        amount = float(row["Amount"])

        if row["Type"] == "Income":

            monthly_income += amount
            income_count += 1

        else:

            monthly_expenses += amount
            expense_count += 1

    net_savings = monthly_income - monthly_expenses

    budget = load_budget()

    remaining_budget = budget - monthly_expenses

    display_month = datetime.strptime(selected_month, "%Y-%m").strftime("%B %Y")

    report_window = tk.Toplevel(root)

    report_window.title("Monthly Summary")

    report_window.geometry(f"{REPORT_WIDTH}x{REPORT_HEIGHT}")

    report_window.resizable(False, False)

    scrollbar = tk.Scrollbar(report_window)

    scrollbar.pack(side="right", fill="y")

    text = tk.Text(report_window, font=REPORT_FONT, yscrollcommand=scrollbar.set)

    text.pack(fill="both", expand=True, padx=10, pady=10)

    scrollbar.config(command=text.yview)

    enable_mousewheel_scrolling(text)

    report = ""

    report += f"Monthly Summary ({display_month})\n"

    report += "=" * 45 + "\n\n"

    report += f"{'Total Income':<25}₹{monthly_income:,.2f}\n"

    report += f"{'Total Expenses':<25}₹{monthly_expenses:,.2f}\n"

    report += f"{'Net Savings':<25}₹{net_savings:,.2f}\n"

    report += "\n"

    report += f"{'Income Transactions':<25}{income_count}\n"

    report += f"{'Expense Transactions':<25}{expense_count}\n"

    report += "\n"

    report += f"{'Budget':<25}₹{budget:,.2f}\n"

    if budget == 0:

        report += f"{'Budget Status':<25}Not Set\n"

    elif remaining_budget >= 0:

        report += (
            f"{'Budget Status':<25}"
            f"₹{remaining_budget:,.2f} Remaining\n"
        )

    else:

        report += (
            f"{'Budget Status':<25}"
            f"₹{abs(remaining_budget):,.2f} Exceeded\n"
        )

    text.insert("1.0", report)

    text.config(state="disabled")


   
def generate_monthly_summary(selector, month_name, year):

    month_number = datetime.strptime(month_name, "%B").month

    selected_month = f"{year}-{month_number:02d}"

    selector.destroy()

    show_monthly_summary(selected_month)



def show_monthly_category_budget_status(selected_month):

    category_budgets = load_category_budgets()

    if not category_budgets:

        messagebox.showinfo(
            "No Budgets",
            "No category budgets have been set."
        )

        return

    category_spending = {}

    transactions = read_transactions()

    for row in transactions:

        if (row["Type"] == "Expense" and row["Date"].startswith(selected_month)):

            category = row["Category"]

            amount = float(row["Amount"])

            category_spending[category] = (category_spending.get(category, 0) + amount)

    display_month = datetime.strptime(selected_month, "%Y-%m").strftime("%B %Y")

    report_window = tk.Toplevel(root)

    report_window.title(f"Category Budget Status ({display_month})")

    report_window.geometry(f"{REPORT_WIDTH}x{REPORT_HEIGHT}")

    all_categories = sorted(set(get_all_categories()) | set(category_budgets.keys()))

    report = (f"Category Budget Status ({display_month})\n")

    report += "=" * 75 + "\n\n"

    report = (
        f"{'Category':<20}"
        f"{'Budget':>15}"
        f"{'Spent':>15}"
        f"{'Status':>25}\n"
    )

    report += "-" * 75 + "\n\n"

    for category in all_categories:

        budget = category_budgets.get(category)

        spent = category_spending.get(category, 0)

        if budget is None:

            budget_text = "Not Set"
            status = "Not Set"

        else:

            budget_text = f"₹{budget:,.2f}"

            if spent < budget:

                status = f"₹{budget - spent:,.2f} Remaining"

            elif spent > budget:

                status = f"₹{spent - budget:,.2f} Exceeded"

            else:

                status = "Limit Reached"

        if spent == 0:

            spent_text = "No Spending"

        else:

            spent_text = f"₹{spent:,.2f}"

        report += (
            f"{category:<20}"
            f"{budget_text:>15}"
            f"{spent_text:>15}"
            f"{status:>25}\n"
        )

    scrollbar = tk.Scrollbar(report_window)

    scrollbar.pack(side="right", fill="y")

    text = tk.Text(report_window, wrap="none", font=REPORT_FONT, yscrollcommand=scrollbar.set)

    text.pack(fill="both", expand=True, padx=10, pady=10)

    scrollbar.config(command=text.yview)

    enable_mousewheel_scrolling(text)

    text.insert(tk.END, report)

    text.config(state="disabled")



def generate_category_budget_status(selector, month_name, year):

    month_number = datetime.strptime(month_name, "%B").month

    selected_month = f"{year}-{month_number:02d}"

    selector.destroy()

    show_monthly_category_budget_status(selected_month)



def show_monthly_trend():

    monthly_expenses = {}

    transactions = read_transactions()

    for row in transactions:

        if row["Type"] == "Expense":

            month = row["Date"][:7]      # YYYY-MM

            amount = float(row["Amount"])

            if month not in monthly_expenses:

                monthly_expenses[month] = 0

            monthly_expenses[month] += amount

    if not monthly_expenses:

        messagebox.showinfo(
            "No Data",
            "No expense data available."
        )
        return

    months = sorted(monthly_expenses.keys())

    start_month = datetime.strptime(months[0], "%Y-%m")
    end_month = datetime.strptime(months[-1], "%Y-%m")

    all_months = []

    current = start_month

    while current <= end_month:

        month = current.strftime("%Y-%m")

        all_months.append(month)

        if current.month == 12:

            current = current.replace(
                year=current.year + 1,
                month=1
            )

        else:

            current = current.replace(
                month=current.month + 1
            )

    expenses = [
        monthly_expenses.get(month, 0)
        for month in all_months
    ]

    plt.figure(figsize=(8, 5))

    display_months = [
        datetime.strptime(month, "%Y-%m").strftime("%b %Y")
        for month in all_months
    ]

    plt.plot(
        display_months,
        expenses,
        marker="o"
    )

    plt.title("Monthly Expense Trend")
    plt.xlabel("Month")
    plt.ylabel("Amount (₹)")

    plt.grid(True)

    plt.show()



def open_export_window(selector, month, year):

    selector.destroy()

    selected_month = (f"{year}-{datetime.strptime(month, '%B').month:02d}")

    export_window = tk.Toplevel(root)

    export_window.title("Export Financial Report")

    export_window.geometry(f"{EXPORT_WIDTH}x{EXPORT_HEIGHT}")

    export_window.resizable(False, False)

    tk.Label(export_window, text="Select Export Format").pack(pady=(20, 10))

    format_combobox = ttk.Combobox(
        export_window,
        values=EXPORT_FORMATS,
        state="readonly",
        width=20
    )

    format_combobox.pack()

    format_combobox.current(0)

    tk.Button(
        export_window,
        text="Export",
        command=lambda: export_selected_format(
            format_combobox,
            export_window,
            selected_month
        )
    ).pack(pady=20)



def export_selected_format(format_combobox, export_window, selected_month):

    selected_format = format_combobox.get()

    export_window.destroy()

    if selected_format == "CSV":

        export_csv(selected_month)

    elif selected_format == "PDF":

        export_pdf(selected_month)

    else:

        export_excel(selected_month)



def export_report():

    open_month_selector(
        "Export Report",
        open_export_window
    )



def export_csv(selected_month):

    file_path = filedialog.asksaveasfilename(
        defaultextension=".csv",
        filetypes=[("CSV Files", "*.csv")],
        title="Save Financial Report"
    )

    if not file_path:

        return

    try:

        with open(file_path, "w", newline="", encoding="utf-8") as file:

            writer = csv.writer(file)

            writer.writerow(CSV_HEADERS)

            transactions = read_transactions()

            for row in transactions:

                if row["Date"].startswith(selected_month):

                    writer.writerow([
                        row["ID"],
                        row["Date"],
                        row["Type"],
                        row["Category"],
                        row["Amount"],
                        row["Description"]
                    ])

        messagebox.showinfo(
            "Success",
            "Financial report exported successfully."
        )

    except Exception as e:

        messagebox.showerror(
            "Export Error",
            f"Unable to export report.\n\n{e}"
        )



def export_pdf(selected_month):

    file_path = filedialog.asksaveasfilename(
        defaultextension=".pdf",
        filetypes=[("PDF Files", "*.pdf")],
        title="Save Financial Report"
    )

    if not file_path:

        return
    
    pdf = SimpleDocTemplate(file_path, pagesize=landscape(A4))

    styles = getSampleStyleSheet()

    title_style = styles["Heading1"]

    normal_style = styles["Normal"]
    
    table_data = [["Date", "Type", "Category", "Amount", "Description"]]

    transactions = read_transactions()

    for row in transactions:

        if not row["Date"].startswith(selected_month):

            continue

        table_data.append([
            row["Date"],
            row["Type"],
            row["Category"],
            f"Rs.{float(row['Amount']):,.2f}",
            row["Description"]
        ])

    month_display = datetime.strptime(selected_month, "%Y-%m").strftime("%B %Y")

    title = Paragraph(f"Personal Finance Report ({month_display})",title_style)

    generated_on = Paragraph(
        f"Generated On: {datetime.now().strftime('%d %B %Y, %I:%M %p')}",
        normal_style
    )

    table = Table(table_data, colWidths=[90, 70, 100, 80, 240], repeatRows=1)

    table.setStyle(

        TableStyle([

            ("BACKGROUND", (0, 0), (-1, 0), colors.grey),

            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),

            ("ALIGN", (0, 0), (-1, -1), "CENTER"),

            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),

            ("BOTTOMPADDING", (0, 0), (-1, 0), 10),

            ("GRID", (0, 0), (-1, -1), 1, colors.black),

            ("BACKGROUND", (0, 1), (-1, -1), colors.beige)

        ])

    )

    try:

        table.splitByRow = True

        pdf.build([title, Spacer(1, 12), generated_on, Spacer(1, 20), table])

        messagebox.showinfo(
            "Success",
            "Financial report exported successfully."
        )

    except Exception as e:

        messagebox.showerror(
            "Export Error",
            f"Unable to export report.\n\n{e}"
        )



def export_excel(selected_month):

    file_path = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Excel Files", "*.xlsx")],
        title="Save Financial Report"
    )

    if not file_path:

        return

    try:

        workbook = Workbook()

        worksheet = workbook.active

        worksheet.title = "Financial Report"

        worksheet.append(CSV_HEADERS)

        for cell in worksheet[1]:

            cell.font = Font(bold=True)

        transactions = read_transactions()

        for row in transactions:

            if not row["Date"].startswith(selected_month):
                
                continue

            worksheet.append([
                row["ID"],
                row["Date"],
                row["Type"],
                row["Category"],
                float(row["Amount"]),
                row["Description"]
            ])

        # Format Amount Column
        for cell in worksheet["E"][1:]:

            cell.number_format = '#,##0.00'

        # Auto-fit Columns
        for column_cells in worksheet.columns:

            max_length = 0

            column = get_column_letter(column_cells[0].column)

            for cell in column_cells:

                try:

                    if len(str(cell.value)) > max_length:

                        max_length = len(str(cell.value))

                except:

                    pass

            worksheet.column_dimensions[column].width = max_length + 2
        
        workbook.save(file_path)

        messagebox.showinfo(
            "Success",
            "Financial report exported successfully."
        )

    except Exception as e:

        messagebox.showerror(
            "Export Error",
            f"Unable to export report.\n\n{e}"
        )



def shortcut_add(event=None):
    add_transaction()

def shortcut_save(event=None):
    save_changes()

def shortcut_delete(event=None):
    delete_transaction()

def shortcut_undo(event=None):
    undo_delete()

def shortcut_edit(event=None):
    edit_transaction()

def shortcut_search(event=None):
    search_entry.focus_set()

def shortcut_clear(event=None):
    clear_inputs()



create_csv_file()

root = tk.Tk()
root.title("Personal Finance Dashboard")
root.geometry(f"{WINDOW_WIDTH}x{WINDOW_HEIGHT}")
root.resizable(True, True)

# Title
title_label = tk.Label(
    root,
    text="Personal Finance Dashboard",
    font=TITLE_FONT
)
title_label.pack(pady=10)

#Summary Frame
summary_frame = tk.Frame(root)

summary_frame.pack(pady=10)

income_label = tk.Label(
    summary_frame,
    text="Total Income: ₹0.00",
    font=LABEL_FONT
)

income_label.grid(row=0, column=0, padx=20)

expense_label = tk.Label(
    summary_frame,
    text="Total Expenses: ₹0.00",
    font=LABEL_FONT
)

expense_label.grid(row=0, column=1, padx=20)

balance_label = tk.Label(
    summary_frame,
    text="Current Balance: ₹0.00",
    font=LABEL_FONT
)

balance_label.grid(row=0, column=2, padx=20)

budget_label = tk.Label(
    summary_frame,
    text="Monthly Budget: ₹0.00",
    font=LABEL_FONT
)

budget_label.grid(row=1, column=0, padx=20)

status_label = tk.Label(
    summary_frame,
    text="Budget Status: Not Set",
    font=LABEL_FONT
)

status_label.grid(row=1, column=2, padx=20)

# Budget Frame
budget_frame = tk.Frame(root)

budget_frame.pack(pady=5)

tk.Label(budget_frame, text="Monthly Budget").grid(row=0, column=0, padx=5)

budget_entry = tk.Entry(budget_frame, width=20)

budget_entry.grid(row=0, column=1, padx=5)

saved_budget = load_budget()

if saved_budget > 0:
    budget_entry.insert(0, str(saved_budget))

# Set Budget Button
budget_button = tk.Button(
    budget_frame,
    text="Set Budget",
    command=save_budget
)

budget_button.grid(row=0, column=2, padx=5)

# Reset Budget Button
reset_budget_button = tk.Button(
    budget_frame,
    text="Reset Budget",
    command=reset_budget
)

reset_budget_button.grid(row=0, column=3, padx=5)

# Category Budgets Button
category_budget_button = tk.Button(
    budget_frame,
    text="Category Budgets",
    command=open_category_budget_window
)

category_budget_button.grid(row=0, column=4, padx=5)

# Input Frame
input_frame = tk.Frame(root)
input_frame.pack(pady=10)

# Date
tk.Label(input_frame, text="Date").grid(row=0, column=0, padx=5, pady=5)

date_entry = DateEntry(
    input_frame,
    width=20,
    date_pattern="yyyy-mm-dd"
)

date_entry.grid(row=0, column=1, padx=5, pady=5)

# Type
tk.Label(input_frame, text="Type").grid(row=1, column=0, padx=5, pady=5)

type_combobox = ttk.Combobox(
    input_frame,
    values=["Expense", "Income"],
    state="readonly",
    width=20
)
type_combobox.grid(row=1, column=1, padx=5, pady=5)
type_combobox.current(0)

# Category
tk.Label(input_frame, text="Category").grid(row=2, column=0, padx=5, pady=5)

category_combobox = ttk.Combobox(
    input_frame,
    values=CATEGORIES,
    state="readonly",
    width=20
)
category_combobox.grid(row=2, column=1, padx=5, pady=5)
category_combobox.current(0)

#Custom category
custom_category_label = tk.Label(input_frame, text="Custom Category")
custom_category_entry = tk.Entry(input_frame, width=23)

custom_category_label.grid(row=5, column=0, padx=5, pady=5)
custom_category_entry.grid(row=5, column=1, padx=5, pady=5)

custom_category_label.grid_remove()
custom_category_entry.grid_remove()

# Amount
tk.Label(input_frame, text="Amount").grid(row=3, column=0, padx=5, pady=5)

amount_entry = tk.Entry(input_frame, width=23)
amount_entry.grid(row=3, column=1, padx=5, pady=5)

# Description
tk.Label(input_frame, text="Description").grid(row=4, column=0, padx=5, pady=5)

description_entry = tk.Entry(input_frame, width=23)
description_entry.grid(row=4, column=1, padx=5, pady=5)

category_combobox.bind(
    "<<ComboboxSelected>>",
    handle_category_change
)

# Add Transaction Button
add_button = tk.Button(
    input_frame,
    text="Add Transaction",
    command=add_transaction
)

add_button.grid(row=6, column=0, columnspan=2, pady=10)

# Search Frame
search_frame = tk.Frame(root)
search_frame.pack(pady=10)

tk.Label(search_frame, text="Search").grid(row=0, column=0, padx=5)

search_entry = tk.Entry(search_frame, width=30)

search_entry.grid(row=0, column=1, padx=5)

search_entry.bind(
    "<KeyRelease>",
    debounce_search
)

# Filter
tk.Label(search_frame, text="Type").grid(row=0, column=2, padx=5)

filter_combobox = ttk.Combobox(
    search_frame,
    values=["All", "Income", "Expense"],
    state="readonly",
    width=15
)

filter_combobox.grid(row=0, column=3,padx=5)
filter_combobox.current(0)

filter_combobox.bind(
    "<<ComboboxSelected>>",
    lambda event: apply_filter()
)

# Reset Filter Button
reset_button = tk.Button(
    search_frame,
    text="Reset Filter",
    command=reset_filter
)

reset_button.grid(row=0, column=4, padx=5)

# Action Frame
action_frame = tk.Frame(root)

action_frame.pack(pady=10)

# Edit Transaction Button
edit_button = tk.Button(
    action_frame,
    text="Edit Selected Transaction",
    command=edit_transaction
)

edit_button.grid(row=0, column=0, padx=5, pady=5)

# Save Changes Button
save_button = tk.Button(
    action_frame,
    text="Save Changes",
    command=save_changes
)

save_button.grid(row=0, column=1, padx=5, pady=5)

# Delete Transaction(s) Button
delete_button = tk.Button(
    action_frame,
    text="Delete Selected Transaction(s)",
    command=delete_transaction
)

delete_button.grid(row=0, column=2, padx=5, pady=5)

# Undo Delete Button
undo_delete_button = tk.Button(
    action_frame,
    text="Undo Delete",
    command=undo_delete,
    state="disabled"
)

undo_delete_button.grid(row=0, column=3, padx=5, pady=5)

# Monthly Expense Breakdown Button
chart_button = tk.Button(
    action_frame,
    text="Monthly Expense Breakdown",
    command=lambda: open_month_selector(
        "Show Expense Breakdown",
        generate_expense_breakdown
    )
)

chart_button.grid(row=1, column=0, padx=5, pady=5)

# Monthly Category-Wise Spending Report 
category_report_button = tk.Button(
    action_frame,
    text="Monthly Category Report",
    command=lambda: open_month_selector(
        "Show Report",
        generate_category_report
    )
)

category_report_button.grid(row=2, column=0, padx=5, pady=5)

# Monthly Summary Button
monthly_summary_button = tk.Button(
    action_frame,
    text="Monthly Summary",
    command=lambda: open_month_selector(
        "Show Summary",
        generate_monthly_summary
    )
)

monthly_summary_button.grid(row=2, column=1, padx=5, pady=5)

# Monhly Category Budget Status Button
category_budget_status_button = tk.Button(
    action_frame,
    text="Monthly Category Budget Status",
    command=lambda: open_month_selector(
        "Show Status",
        generate_category_budget_status
    )
)

category_budget_status_button.grid(row=2, column=2, padx=5, pady=5)

# Monthly Expense Trend Button
trend_button = tk.Button(
    action_frame,
    text="Monthly Trend",
    command=show_monthly_trend
)

trend_button.grid(row=1, column=1, padx=5, pady=5)

# Export button
export_button = tk.Button(
    action_frame,
    text="Export Financial Report",
    command=export_report
)

export_button.grid(row=1, column=2, padx=5, pady=5)

# Table Frame
table_frame = tk.Frame(root)

table_frame.pack(fill="both", expand=True, padx=20, pady=20)

# Transaction Table
columns = ("ID", "Date", "Type", "Category", "Amount", "Description")

tree = ttk.Treeview(
    table_frame,
    columns=columns,
    show="headings",
    selectmode="extended"
)

for column in columns:

    if column in ("Date", "Amount", "Category"):

        tree.heading(
            column,
            text=column,
            command=lambda c=column: sort_treeview(c)
        )

    else:

        tree.heading(
            column,
            text=column
        )

    if column == "ID":

        tree.column(column, width=0, stretch=False)

    elif column == "Description":

        tree.column(column, width=350, minwidth=250, anchor="w", stretch=True)

    elif column == "Category":

        tree.column(column, width=180, minwidth=140, anchor="center", stretch=True)

    elif column == "Amount":

        tree.column(column, width=120, minwidth=100, anchor="e", stretch=False)

    else:

        tree.column(column, width=140, minwidth=100, anchor="center", stretch=True)

update_sort_headers()

tree.tag_configure(
    "evenrow",
    background=EVEN_ROW_COLOR
)

tree.tag_configure(
    "oddrow",
    background=ODD_ROW_COLOR
)

# Vertical Scrollbar
scrollbar_y = ttk.Scrollbar(
    table_frame,
    orient="vertical",
    command=tree.yview
)

tree.configure(yscrollcommand=scrollbar_y.set)

enable_mousewheel_scrolling(tree)

# Horizontal Scrollbar
scrollbar_x = ttk.Scrollbar(
    table_frame,
    orient="horizontal",
    command=tree.xview
)

tree.configure(xscrollcommand=scrollbar_x.set)

# Pack Widgets
scrollbar_y.pack(side="right", fill="y")

scrollbar_x.pack(side="bottom", fill="x")

tree.pack(side="left", fill="both", expand=True)



load_transactions()

update_summary()

root.bind("<Return>", shortcut_add)

root.bind("<Control-s>", shortcut_save)

root.bind("<Delete>", shortcut_delete)

root.bind("<Control-z>", shortcut_undo)

root.bind("<Control-e>", shortcut_edit)

root.bind("<Control-f>", shortcut_search)

root.bind("<Escape>", shortcut_clear)

root.mainloop()
